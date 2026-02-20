"""
Generate Excel test fixtures for excel-anonymizer integration tests.

Run this script to (re)create the fixture files:
    cd /Users/danielbutler/code/excel-anon
    uv run python tests/create_fixtures.py

Fixtures:
    tests/fixtures/Dec-24 Revenue Report.xlsx  - Multi-sheet financial report
    tests/fixtures/2024-Q4 Team Roster.xlsx    - Single-sheet employee data
"""
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


FIXTURES_DIR = Path(__file__).parent / "fixtures"


def create_revenue_report():
    """
    Multi-sheet financial report: Dec-24 Revenue Report.xlsx

    Sheets:
        Summary    - Monthly totals per client (revenue, cost, GM, GM%)
        Details    - Line-item project data (project names, managers, descriptions)
        Team       - Sales team roster (names, emails, phone numbers)

    Tricky aspects:
        - GM = Revenue - Cost (PreserveRelationshipRule)
        - GM% = GM / Revenue * 100 (PreserveRelationshipRule)
        - Same people appear across Summary (as managers) and Team (as employees)
        - Client names appear in both Summary and Details
        - Mix of entity types: PERSON, ORGANIZATION, PROJECT_NAME, EMAIL_ADDRESS, PHONE_NUMBER
        - Null values in optional Description column
    """
    wb = Workbook()

    # --- Sheet 1: Summary ---
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary_data = [
        ["Client", "Account Manager", "Revenue", "Cost", "Gross Margin", "GM%"],
        ["Northgate Industries", "Sarah Chen",    142500.00, 98750.00, 43750.00, 30.70],
        ["Apex Solutions LLC",  "Marcus Webb",    87300.00,  61200.00, 26100.00, 29.90],
        ["Riverfront Corp",     "Sarah Chen",     210000.00, 147000.00, 63000.00, 30.00],
        ["Pinnacle Group",      "Jordan Hayes",   55800.00,  41200.00, 14600.00, 26.16],
        ["Coastal Dynamics",    "Marcus Webb",    178400.00, 124100.00, 54300.00, 30.44],
        ["Summit Partners",     "Jordan Hayes",   93200.00,  65400.00, 27800.00, 29.83],
        ["Meridian Tech",       "Sarah Chen",     315000.00, 220500.00, 94500.00, 30.00],
        ["Harborview Systems",  "Marcus Webb",    44100.00,  32900.00, 11200.00, 25.40],
        ["Irongate Ventures",   "Jordan Hayes",   127600.00, 89320.00, 38280.00, 30.00],
        ["Clearwater Group",    "Sarah Chen",     68900.00,  49400.00, 19500.00, 28.30],
        ["Total", "",           1322800.00, 930770.00, 392030.00, 29.64],
    ]

    for row in summary_data:
        ws_summary.append(row)

    # Header styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for cell in ws_summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Bold the totals row
    for cell in ws_summary[len(summary_data) + 1]:
        cell.font = Font(bold=True)

    # Column widths
    ws_summary.column_dimensions["A"].width = 22
    ws_summary.column_dimensions["B"].width = 18
    for col in ["C", "D", "E", "F"]:
        ws_summary.column_dimensions[col].width = 16

    # --- Sheet 2: Details ---
    ws_details = wb.create_sheet("Details")

    details_data = [
        ["Project Name", "Client", "Project Manager", "Description", "Start Date", "End Date", "Billed", "Expenses", "Net"],
        ["Cloud Migration Phase 1",     "Northgate Industries",  "Sarah Chen",    "Infrastructure lift-and-shift to AWS",                     "2024-10-01", "2024-12-31", 85000.00,  60000.00, 25000.00],
        ["ERP Integration",             "Northgate Industries",  "Daniel Park",   "SAP integration with legacy billing system",                "2024-11-01", "2024-12-31", 57500.00,  38750.00, 18750.00],
        ["Sales Analytics Dashboard",   "Apex Solutions LLC",    "Marcus Webb",   "Power BI dashboard for pipeline visibility",                "2024-09-15", "2024-12-31", 87300.00,  61200.00, 26100.00],
        ["Digital Transformation",      "Riverfront Corp",       "Sarah Chen",    "Full-stack modernization of customer portal",               "2024-07-01", "2024-12-31", 130000.00, 91000.00, 39000.00],
        ["API Gateway Rollout",         "Riverfront Corp",       "Priya Nair",    None,                                                        "2024-10-15", "2024-12-31", 80000.00,  56000.00, 24000.00],
        ["Compliance Audit Prep",       "Pinnacle Group",        "Jordan Hayes",  "SOC 2 Type II readiness assessment",                        "2024-11-01", "2024-12-15", 55800.00,  41200.00, 14600.00],
        ["Marketing Automation",        "Coastal Dynamics",      "Marcus Webb",   "HubSpot implementation and CRM migration",                  "2024-08-01", "2024-12-31", 98400.00,  68100.00, 30300.00],
        ["Data Warehouse Build",        "Coastal Dynamics",      "Priya Nair",    "Snowflake DW with dbt transformation layer",                "2024-10-01", "2024-12-31", 80000.00,  56000.00, 24000.00],
        ["Security Assessment",         "Summit Partners",       "Jordan Hayes",  None,                                                        "2024-11-15", "2024-12-31", 93200.00,  65400.00, 27800.00],
        ["Platform Engineering",        "Meridian Tech",         "Sarah Chen",    "Kubernetes cluster build-out and CI/CD pipeline",           "2024-06-01", "2024-12-31", 195000.00, 136500.00, 58500.00],
        ["AI Prototype",                "Meridian Tech",         "Daniel Park",   "LLM-powered document processing proof of concept",          "2024-10-01", "2024-12-31", 120000.00, 84000.00, 36000.00],
        ["Infrastructure Review",       "Harborview Systems",    "Marcus Webb",   None,                                                        "2024-11-01", "2024-12-31", 44100.00,  32900.00, 11200.00],
        ["Growth Strategy Roadmap",     "Irongate Ventures",     "Jordan Hayes",  "Go-to-market strategy for Series B expansion",              "2024-09-01", "2024-12-31", 127600.00, 89320.00, 38280.00],
        ["Customer 360 Implementation", "Clearwater Group",      "Sarah Chen",    "Salesforce implementation with custom CPQ",                 "2024-10-15", "2024-12-31", 68900.00,  49400.00, 19500.00],
    ]

    for row in details_data:
        ws_details.append(row)

    for cell in ws_details[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws_details.column_dimensions["A"].width = 30
    ws_details.column_dimensions["B"].width = 22
    ws_details.column_dimensions["C"].width = 18
    ws_details.column_dimensions["D"].width = 45
    for col in ["E", "F"]:
        ws_details.column_dimensions[col].width = 14
    for col in ["G", "H", "I"]:
        ws_details.column_dimensions[col].width = 14

    # --- Sheet 3: Team ---
    ws_team = wb.create_sheet("Team")

    team_data = [
        ["Name", "Title", "Email", "Phone", "Office"],
        ["Sarah Chen",    "Principal Consultant",   "s.chen@example-consulting.com",    "555-0142", "Seattle, WA"],
        ["Marcus Webb",   "Senior Consultant",      "m.webb@example-consulting.com",    "555-0198", "Portland, OR"],
        ["Jordan Hayes",  "Consultant",             "j.hayes@example-consulting.com",   "555-0231", "San Francisco, CA"],
        ["Daniel Park",   "Associate Consultant",   "d.park@example-consulting.com",    "555-0187", "Seattle, WA"],
        ["Priya Nair",    "Senior Consultant",      "p.nair@example-consulting.com",    "555-0264", "San Jose, CA"],
    ]

    for row in team_data:
        ws_team.append(row)

    for cell in ws_team[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws_team.column_dimensions["A"].width = 18
    ws_team.column_dimensions["B"].width = 24
    ws_team.column_dimensions["C"].width = 34
    ws_team.column_dimensions["D"].width = 14
    ws_team.column_dimensions["E"].width = 20

    output_path = FIXTURES_DIR / "Dec-24 Revenue Report.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")


def create_team_roster():
    """
    Single-sheet employee data: 2024-Q4 Team Roster.xlsx

    Tricky aspects:
        - Separate first name and last name columns (PERSON_FIRST_NAME, PERSON_LAST_NAME)
        - Full name column as well (PERSON)
        - Mixed entity types on the same row
        - Some employees have no phone (nulls)
        - Salary column to anonymize with variance
        - Annual bonus = Salary * Bonus % (PreserveRelationshipRule)
        - Dates (hire date) to preserve
        - Location column (LOCATION)
    """
    data = {
        "Employee ID": [1001, 1002, 1003, 1004, 1005, 1006, 1007, 1008, 1009, 1010,
                        1011, 1012, 1013, 1014, 1015],
        "First Name":  ["James", "Linda", "Robert", "Patricia", "Michael",
                        "Barbara", "William", "Susan", "David", "Jessica",
                        "Richard", "Karen", "Joseph", "Nancy", "Thomas"],
        "Last Name":   ["Morrison", "Fitzgerald", "Okafor", "Sullivan", "Hernandez",
                        "Whitfield", "Nakamura", "Delgado", "Patel", "Thornton",
                        "Beaumont", "Ashworth", "Reyes", "Yamamoto", "Obrien"],
        "Full Name":   ["James Morrison", "Linda Fitzgerald", "Robert Okafor",
                        "Patricia Sullivan", "Michael Hernandez", "Barbara Whitfield",
                        "William Nakamura", "Susan Delgado", "David Patel",
                        "Jessica Thornton", "Richard Beaumont", "Karen Ashworth",
                        "Joseph Reyes", "Nancy Yamamoto", "Thomas Obrien"],
        "Department":  ["Engineering", "Marketing", "Engineering", "Finance", "Sales",
                        "Engineering", "Product", "HR", "Engineering", "Sales",
                        "Finance", "Marketing", "Sales", "Product", "Engineering"],
        "Title":       ["Staff Engineer", "Marketing Director", "Senior Engineer", "Controller",
                        "Account Executive", "Principal Engineer", "Product Manager",
                        "HR Business Partner", "Senior Engineer", "Sales Manager",
                        "VP Finance", "Content Strategist", "Account Executive",
                        "Senior PM", "Engineer II"],
        "Email":       ["j.morrison@example.com", "l.fitzgerald@example.com",
                        "r.okafor@example.com", "p.sullivan@example.com",
                        "m.hernandez@example.com", "b.whitfield@example.com",
                        "w.nakamura@example.com", "s.delgado@example.com",
                        "d.patel@example.com", "j.thornton@example.com",
                        "r.beaumont@example.com", "k.ashworth@example.com",
                        "j.reyes@example.com", "n.yamamoto@example.com",
                        "t.obrien@example.com"],
        "Phone":       ["555-1001", "555-1002", None, "555-1004", "555-1005",
                        "555-1006", None, "555-1008", "555-1009", "555-1010",
                        "555-1011", None, "555-1013", "555-1014", "555-1015"],
        "Location":    ["Austin, TX", "New York, NY", "Chicago, IL", "Boston, MA", "Dallas, TX",
                        "Austin, TX", "San Francisco, CA", "Chicago, IL", "Seattle, WA", "Dallas, TX",
                        "Boston, MA", "New York, NY", "Phoenix, AZ", "San Francisco, CA", "Austin, TX"],
        "Hire Date":   ["2019-03-15", "2017-08-01", "2021-06-14", "2016-01-10", "2022-04-03",
                        "2018-11-27", "2020-09-08", "2023-02-20", "2021-07-19", "2022-10-31",
                        "2015-05-12", "2019-12-02", "2023-08-14", "2020-03-30", "2022-06-06"],
        "Base Salary": [145000, 128000, 138000, 115000, 95000,
                        162000, 132000, 108000, 141000, 118000,
                        195000, 105000, 92000, 129000, 112000],
        "Bonus %":     [15, 12, 15, 10, 20,
                        18, 12, 8, 15, 20,
                        20, 10, 20, 12, 10],
        "Annual Bonus": [21750, 15360, 20700, 11500, 19000,
                         29160, 15840, 8640, 21150, 23600,
                         39000, 10500, 18400, 15480, 11200],
    }

    df = pd.DataFrame(data)

    output_path = FIXTURES_DIR / "2024-Q4 Team Roster.xlsx"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Roster", index=False)

        ws = writer.sheets["Roster"]
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        col_widths = [12, 12, 16, 20, 14, 22, 28, 12, 18, 12, 14, 10, 14]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = width

    print(f"Created: {output_path}")


if __name__ == "__main__":
    FIXTURES_DIR.mkdir(parents=True, exist_ok=True)
    create_revenue_report()
    create_team_roster()
    print("Done.")
